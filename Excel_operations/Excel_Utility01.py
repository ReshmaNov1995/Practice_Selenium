import openpyxl
from openpyxl.styles import PatternFill


def max_row(excel_path):
    excel_workbook=openpyxl.load_workbook(excel_path)
    excel_worksheet=excel_workbook.active
    excel_max_row=excel_worksheet.max_row
    return excel_max_row

def max_column(excel_path):
    excel_workbook=openpyxl.load_workbook(excel_path)
    excel_worksheet=excel_workbook.active
    excel_max_column=excel_worksheet.max_column
    return excel_max_column


def read_excel(excel_path,row_no,column_no):
    excel_workbook=openpyxl.load_workbook(excel_path)
    excel_worksheet=excel_workbook.active
    excel_read=excel_worksheet.cell(row_no,column_no).value
    return excel_read

def write_excel(excel_path,row_no,column_no,data):
    excel_workbook=openpyxl.load_workbook(excel_path)
    excel_worksheet=excel_workbook.active
    excel_write=excel_worksheet.cell(row_no,column_no).value=data
    excel_workbook.save(excel_path)


def green_excel(excel_path,row_no,column_no):
    excel_workbook=openpyxl.load_workbook(excel_path)
    excel_worksheet=excel_workbook.active
    green_color=PatternFill(start_color="60d212",end_color="60d212",fill_type="solid")
    excel_write=excel_worksheet.cell(row_no,column_no).fill=green_color
    excel_workbook.save(excel_path)

def red_excel(excel_path,row_no,column_no):
    excel_workbook=openpyxl.load_workbook(excel_path)
    excel_worksheet=excel_workbook.active
    red_color=PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")
    excel_write=excel_worksheet.cell(row_no,column_no).fill=red_color
    excel_workbook.save(excel_path)