import openpyxl

file = "E:\\Automation\\Data\\WriteWorksheet.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
# sheet = workbook.active # If there is only  sheet

# Same Data
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Hi"
# workbook.save(file)

# Multiple Data
sheet.cell(1,1).value="ID"
sheet.cell(1,2).value="Name"
sheet.cell(1,3).value="Location"

sheet.cell(2,1).value="01"
sheet.cell(2,2).value="Reshma"
sheet.cell(2,3).value="Cuddalore"

sheet.cell(3,1).value="02"
sheet.cell(3,2).value="Murali"
sheet.cell(3,3).value="Mannargudi"

sheet.cell(4,1).value="03"
sheet.cell(4,2).value="Keerthana"
sheet.cell(4,3).value="Madurai"

workbook.save(file)