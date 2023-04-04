import openpyxl

# File->workbook->sheets->rows->columns/cells

file = "E:\\Automation\\Data\\DataWorksheet.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"] # sheet name

rows = sheet.max_row # count no.of.rows
columns = sheet.max_column # count no.of.columns

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value, end='   ') # this end='  ' will write the data with tab space instead of writing in next line.
    print() # Seperate the Data to next line
