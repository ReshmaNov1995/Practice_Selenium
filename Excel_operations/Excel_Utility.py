import openpyxl

from openpyxl.styles import PatternFill # Fill the color

def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column

def readData(file, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum, columnnum).value

def writeData(file, sheetname, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum, columnnum).value=data
    workbook.save(file)

def fillGreencolor(file, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenFill = PatternFill(start_color="60b212", end_color="60b212", fill_type="solid")
    sheet.cell(rownum, columnnum).fill=greenFill
    workbook.save(file)

def fillRedcolor(file, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redFill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
    sheet.cell(rownum, columnnum).fill=redFill
    workbook.save(file)

