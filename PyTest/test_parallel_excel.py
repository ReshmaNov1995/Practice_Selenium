import time
# import psutil
import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


sec=4
def xldata():
    excel_path = "E:\Myn\kk.xlsx"

    open_workbook = openpyxl.load_workbook(excel_path)
    open_worksheet = open_workbook["Test Case"]
    max_row = open_worksheet.max_row
    max_column = open_worksheet.max_column
    kkk = []

    print(max_row, max_column)
    for i in range(2, max_row + 1):
        kk = []
        for j in range(1, max_column + 1):
            data = open_worksheet.cell(i, j).value
            kk.insert(j, data) #index, data

        print(kk)
        kkk.append(kk)
    return kkk


@pytest.mark.parametrize("var1,var2", xldata())
def test_eval1(var1, var2):

    driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
    # driver = webdriver.Chrome(executable_path="E:\KRN Files\Python\Selenium\Youtube\Browser Driver\chromedriver.exe")
    driver.get("http://62.171.183.83:9092/QuaLISWeb/#/login")
    time.sleep(4)
    driver.find_element(By.XPATH, "//*[@name='sloginid']").send_keys(var1)
    time.sleep(4)
    driver.find_element(By.XPATH, "//*[@name='spassword']").send_keys(var2)
    time.sleep(4)
    driver.find_element(By.XPATH, "//*[text()='Login']").click()
    # print('\nThe CPU usage in ', sec, 'second is:', psutil.cpu_percent(1), "%")
    time.sleep(8)

